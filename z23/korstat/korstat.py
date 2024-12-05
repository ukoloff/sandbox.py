#
# Перенос записей из C:\Paket\Baza\KorStat.dbf в \\kzkserv\Data\Baza_23\stat\stat.XLSX
#
import os
import glob
import shutil
import json
import hashlib
from datetime import date, datetime
import openpyxl
from dbfread import DBF
import openpyxl.cell
import openpyxl.styles

src = r"C:\Paket\Baza\KorStat.dbf"
# dst = r"\\kzkserv\Data\Baza_23\stat\stat.xlsx"
dst = r"C:\Paket\stat\DBFs\stat.xlsx"

# dst = os.path.join(__file__, "../../../tmp/stat.xlsx")


# https://stackoverflow.com/a/22238613/6127481
def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""

    if isinstance(obj, (datetime, date)):
        return datetime(*obj.timetuple()[:6]).isoformat()
    raise TypeError("Type %s not serializable" % type(obj))


def f2i(x):
    """
    Convert float to int if it is int

    Excel tends to do so
    """
    if isinstance(x, float) and int(x) == x:
        return int(x)
    return x


def hash(data):
    return hashlib.sha256(
        json.dumps(
            [f2i(it) for it in data], ensure_ascii=False, default=json_serial
        ).encode("utf8")
    ).hexdigest()


if not os.path.isfile(src):
    print("File not found: ", src)
    exit()

sources = [src] + glob.glob(os.path.join(os.path.dirname(dst), "*.dbf"))

wbo = openpyxl.Workbook(write_only=True)
wbo.iso_dates = True
wso = wbo.create_sheet()
wso.title = "korStat"
wso.freeze_panes = "A2"

headerHash = None
rowHashes = set()


def add(data, header=False):
    if not header:
        wso.append(data)
        return
    row = []
    font = openpyxl.styles.Font(bold=True)
    align = openpyxl.styles.Alignment(horizontal="center")
    for it in data:
        cell = openpyxl.cell.WriteOnlyCell(wso, value=it)
        cell.font = font
        cell.alignment = align
        row.append(cell)
    wso.append(row)


if os.path.isfile(dst):
    wbi = openpyxl.load_workbook(dst, read_only=True, data_only=True)
    wsi = wbi.active
    for row in wsi:
        data = [cell.value for cell in row]
        add(data, not headerHash)
        h = hash(data)
        if not headerHash:
            headerHash = h
        else:
            rowHashes.add(h)
    wbi.close()

for infile in sources:
    data = DBF(infile, encoding="cp866")
    first = True
    for row in data:
        if first:
            first = False
            hdr = list(row.keys())
            if headerHash:
                if headerHash and hash(hdr) != headerHash:
                    print("Несоответствие списка полей:", *hdr)
            else:
                add(hdr, True)
                headerHash = hash(hdr)
        row = list(row.values())
        h = hash(row)
        if h in rowHashes:
            continue
        rowHashes.add(h)
        add(row)

wbo.save(dst)

try:
    shutil.copy2(dst, os.path.dirname(os.path.dirname(dst)))
except:  # noqa: E722
    print("Cannot copy:", dst)
